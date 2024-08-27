from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, coordinate_to_tuple

## 입력 파일 로드
input_file_path = "./inputs/input_test_01.xlsx"
wb = load_workbook(filename = input_file_path)

## 시트명 확인 및 대상시트 선택
sheetnames_list = wb.sheetnames
print(f"Sheets List : {sheetnames_list}")
ws_name = sheetnames_list[0]
print(f"Target Sheet Name : {ws_name}")
ws = wb[ws_name]

## 대상시트 입력정보 딕셔너리로 저장
data = dict()
for row in ws.iter_rows():
    for idx, cell in enumerate(row):
        if idx == 0:
            key = cell.value
        else:
            value = cell.value
            coordinate = tuple((cell.row, cell.column))
            temp_data = value, coordinate
    data[key] = temp_data
print(data)  #{'가로': (100, (1, 2)), '세로': (50, (2, 2)), '높이': (30, (3, 2))}

가로값 = data["가로"][0] 
가로값좌표 = data["가로"][1]  # (1, 2)
print(f"가로값: {가로값}, 가로좌표: {가로값좌표}")


## 면적, 부피 계산 결과를 수식 포함 엑셀에 저장하기 
# 가로 = f"{get_column_letter(data['가로'][1][1])}{data['가로'][1][0]}"
# 세로 = f"{get_column_letter(data['세로'][1][1])}{data['세로'][1][0]}"
# 높이 = f"{get_column_letter(data['높이'][1][1])}{data['높이'][1][0]}"

def get_coord_value(key_name):   # 함수화 : 가로세로 숫자 인덱스를 "A1", "B1" 형태로 좌표값으로 변경
    return f"{get_column_letter(data[key_name][1][1])}{data[key_name][1][0]}"

가로 = get_coord_value("가로")
세로 = get_coord_value("세로")
높이 = get_coord_value("높이")

면적계산식 = f"={가로} * {세로}"
부피계산식 = f"={가로} * {세로} * {높이}"
print(f"면적: {면적계산식}")
print(f"부피: {부피계산식}")


## 결과파일 저장
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].bestFit = True # 칼럼폭 조정
ws.column_dimensions['B'].width = 30  # 칼럼폭 조정

ws.cell(10, 1).value = "면적"
ws.cell(10, 2).value = 면적계산식
ws.cell(11, 1).value = "부피"
ws.cell(11, 2).value = 부피계산식

wb.save("./outputs/output_test_01.xlsx")
